#!/usr/bin/env python3
"""
Duplicate entity analysis for test_flippro_hostke.

Finds entities with similar names using fuzzy matching, then reports
appointments, insurance mappings, and branches for each duplicate group.

Always writes an Excel workbook alongside terminal output.

Usage:
    python duplicate_entities_analysis.py [--threshold 85] [--dsn "dbname=..."] [--xlsx out.xlsx]
    python duplicate_entities_analysis.py --dry-run
    python duplicate_entities_analysis.py --config config.yaml
"""

import argparse
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from itertools import combinations
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import psycopg2
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
from tabulate import tabulate
from tqdm import tqdm

load_dotenv()

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('duplicate_analysis.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Fuzzy scoring
# ---------------------------------------------------------------------------

# Compiled once at import time
_AMPERSAND_RE = re.compile(r"\s*&\s*")
_PUNCT_RE     = re.compile(r"[^\w\s]")
_SPACE_RE     = re.compile(r"\s+")
# Common legal suffixes that differ between duplicate registrations
_SUFFIX_RE    = re.compile(
    r"\s*\b(ltd|limited|llp|llc|co|company|inc|plc|pvt|pty|group|associates?)\b\.?\s*$",
    re.IGNORECASE,
)


def _normalize(s: str) -> str:
    """Lowercase, replace & with 'and', strip punctuation, collapse spaces."""
    s = _AMPERSAND_RE.sub(" and ", s)
    s = _PUNCT_RE.sub(" ", s)
    return _SPACE_RE.sub(" ", s).strip().lower()


def _strip_suffix(s: str) -> str:
    """Remove trailing legal suffix from a normalised string."""
    return _SUFFIX_RE.sub("", s).strip()


def similarity(a: str, b: str) -> float:
    """
    Returns the best of four scores:
      - ratio and token_sort_ratio on normalised strings
        (catches spacing / punctuation / word-order variants)
      - ratio and token_sort_ratio on suffix-stripped strings
        (catches 'ABC LTD' vs 'ABC LIMITED' vs 'ABC')

    Deliberately excludes partial_ratio and token_set_ratio to avoid
    false positives where unrelated entities share industry keywords
    like 'GARAGE LTD'.
    """
    if not a or not b:
        return 0.0
    an, bn   = _normalize(a),        _normalize(b)
    as_, bs_ = _strip_suffix(an),    _strip_suffix(bn)
    return max(
        fuzz.ratio(an, bn),
        fuzz.token_sort_ratio(an, bn),
        fuzz.ratio(as_, bs_),
        fuzz.token_sort_ratio(as_, bs_),
    )


def is_similar_name(name1: str, name2: str, threshold: int) -> bool:
    """Check if two names are similar above threshold."""
    return similarity(name1, name2) >= threshold


# Emails assigned to on-the-fly entity creation — not a real business email,
# so a shared address here does NOT indicate a duplicate.
_GARBAGE_EMAILS = frozenset({
    "development@innovexsolutions.co.ke",
})


def _real_email(e: dict) -> str:
    addr = (e.get("email") or "").strip().lower()
    return addr if addr and addr not in _GARBAGE_EMAILS else ""


def email_match(a: dict, b: dict) -> bool:
    ea, eb = _real_email(a), _real_email(b)
    return bool(ea and ea == eb)


def entity_similarity(a: dict, b: dict) -> float:
    """Name similarity, boosted to 100 when both entities share a real email."""
    if email_match(a, b):
        return 100.0
    return similarity(a["name"], b["name"])


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

def load_config(config_path: Optional[str] = None) -> dict:
    """Load configuration from YAML file if provided."""
    default_config = {
        'threshold': 85,
        'dsn': 'dbname=test_flippro_hostke',
        'block_size': 100,
        'output_dir': '.',
        'excel_prefix': 'duplicate_report'
    }
    
    if config_path and Path(config_path).exists():
        try:
            with open(config_path) as f:
                user_config = yaml.safe_load(f)
                default_config.update(user_config)
                logger.info(f"Loaded configuration from {config_path}")
        except Exception as e:
            logger.warning(f"Failed to load config {config_path}: {e}")
    
    return default_config


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def _safe_dsn(dsn: str) -> str:
    return re.sub(r"(password\s*=\s*)\S+", r"\1***", dsn, flags=re.IGNORECASE)


def connect(dsn: str):
    """Connect to PostgreSQL database."""
    try:
        conn = psycopg2.connect(dsn)
        logger.info(f"Connected to database: {_safe_dsn(dsn)}")
        return conn
    except psycopg2.OperationalError as e:
        logger.error(f"Cannot connect to database: {e}")
        sys.exit(1)


def fetch(cur, sql: str, params=None) -> list[dict]:
    """Execute query and return results as list of dicts."""
    try:
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        results = [dict(zip(cols, row)) for row in cur.fetchall()]
        logger.debug(f"Query returned {len(results)} rows")
        return results
    except Exception as e:
        logger.error(f"Query failed: {e}")
        logger.error(f"SQL: {sql}")
        raise


# ---------------------------------------------------------------------------
# Data validation
# ---------------------------------------------------------------------------

def validate_entities(entities: List[dict]) -> bool:
    """Validate entity data before processing."""
    issues = []
    for e in entities:
        if not e.get('name'):
            issues.append(f"Entity ID {e.get('id', 'unknown')} has no name")
        if not e.get('id'):
            issues.append("Entity found without ID")
    
    if issues:
        logger.warning(f"Found {len(issues)} data quality issues:")
        for issue in issues[:10]:  # Show first 10
            logger.warning(f"  - {issue}")
        if len(issues) > 10:
            logger.warning(f"  ... and {len(issues) - 10} more")
        return False
    
    logger.info("Entity validation passed")
    return True


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_entities(cur) -> list[dict]:
    """Load all entities with their types."""
    logger.info("Loading entities...")
    entities = fetch(cur, """
        SELECT e.id, e.name, e.email, e.phone_number, e.status,
               e.client_key, e.is_onboarded, e.v_one_id,
               e.physical_address,
               et.name AS entity_type
        FROM entities e
        LEFT JOIN entity_type et ON et.id = e.entity_type_id
        WHERE e.name IS NOT NULL AND e.name != ''
        ORDER BY e.name
    """)
    
    validate_entities(entities)
    logger.info(f"Loaded {len(entities)} entities")
    return entities


def load_appointments(cur) -> dict[str, dict]:
    """Load appointment statistics for service providers."""
    logger.info("Loading appointments...")
    rows = fetch(cur, """
        SELECT service_provider_id::text AS eid,
               COUNT(*)                  AS total,
               COUNT(*) FILTER (WHERE task_status = 'COMPLETED')              AS completed,
               COUNT(*) FILTER (WHERE task_status = 'PENDING')                AS pending,
               COUNT(*) FILTER (WHERE task_status NOT IN ('COMPLETED','PENDING')) AS other
        FROM appointments
        WHERE service_provider_id IS NOT NULL
        GROUP BY service_provider_id
    """)
    result = {r["eid"]: r for r in rows}
    logger.info(f"Loaded appointments for {len(result)} entities")
    return result


def load_insurance_mappings(cur) -> dict[str, list[dict]]:
    """Load insurance mappings for entities."""
    logger.info("Loading insurance mappings...")
    rows = fetch(cur, """
        SELECT ete.service_provider_id::text AS sp_id,
               ete.id::text                  AS mapping_id,
               ete.status                    AS mapping_status,
               ins.name                      AS insurance_name,
               ins.id::text                  AS insurance_id,
               ete.can_book_valuation,
               ete.markup,
               ete.vatable
        FROM entity_to_entity ete
        JOIN entities ins ON ins.id = ete.insurance_id
        ORDER BY ins.name
    """)
    result = defaultdict(list)
    for r in rows:
        result[r["sp_id"]].append(r)
    logger.info(f"Loaded mappings for {len(result)} entities")
    return result


def load_branches(cur) -> dict[str, list[dict]]:
    """Load branch information for entities."""
    logger.info("Loading branches...")
    rows = fetch(cur, """
        SELECT eb.entity_id::text AS eid,
               eb.id::text        AS branch_id,
               eb.email, eb.phone_number, eb.physical_address,
               eb.status,
               aa.name            AS admin_area
        FROM entity_branches eb
        LEFT JOIN admin_areas aa ON aa.id = eb.administrative_area_id
        ORDER BY eb.status
    """)
    result = defaultdict(list)
    for r in rows:
        result[r["eid"]].append(r)
    logger.info(f"Loaded branches for {len(result)} entities")
    return result


# ---------------------------------------------------------------------------
# Duplicate grouping (optimized)
# ---------------------------------------------------------------------------

def find_duplicate_groups_optimized(
    entities: list[dict],
    threshold: int,
    block_size: int = 3,   # kept in signature for config compat, no longer used
) -> list[list[dict]]:
    """
    Complete-linkage clustering over a full O(n²) pair scan.

    An entity joins a group only if it meets the threshold with every existing
    member — not just one chain link. This prevents 'A≈B and B≈C at 60%'
    from pulling unrelated A and C into the same group.
    Pairs are processed strongest-first so high-confidence pairs form group cores.
    """
    n = len(entities)
    logger.info(f"Finding duplicates among {n} entities (threshold: {threshold}%)")

    if n < 2:
        return []

    # Collect all pairs above threshold with their scores.
    sim_pairs: dict[tuple, float] = {}
    total = n * (n - 1) // 2
    with tqdm(total=total, desc="Comparing entities", unit="pair") as pbar:
        for i, j in combinations(range(n), 2):
            s = entity_similarity(entities[i], entities[j])
            if s >= threshold:
                sim_pairs[(i, j)] = s  # i < j always
            pbar.update(1)

    # Complete-linkage clustering — process strongest pairs first.
    group_id: dict[int, int] = {}
    members:  dict[int, set] = {}
    next_gid  = 0

    def all_connected(idx: int, grp: set) -> bool:
        return all((min(idx, m), max(idx, m)) in sim_pairs for m in grp)

    for (i, j) in sorted(sim_pairs, key=lambda p: -sim_pairs[p]):
        gi, gj = group_id.get(i), group_id.get(j)

        if gi is None and gj is None:
            members[next_gid] = {i, j}
            group_id[i] = group_id[j] = next_gid
            next_gid += 1

        elif gi is None:
            if all_connected(i, members[gj]):
                members[gj].add(i)
                group_id[i] = gj

        elif gj is None:
            if all_connected(j, members[gi]):
                members[gi].add(j)
                group_id[j] = gi

        elif gi != gj:
            if all((min(a, b), max(a, b)) in sim_pairs
                   for a in members[gi] for b in members[gj]):
                keep, drop = (gi, gj) if len(members[gi]) >= len(members[gj]) else (gj, gi)
                for m in members[drop]:
                    group_id[m] = keep
                members[keep].update(members.pop(drop))

    result = [
        [entities[i] for i in sorted(grp)]
        for grp in members.values()
        if len(grp) >= 2
    ]
    logger.info(f"Found {len(result)} duplicate groups")
    return result


# Legacy function for backward compatibility
def find_duplicate_groups(entities: list[dict], threshold: int) -> list[list[dict]]:
    """Original duplicate finding logic (maintained for comparison)."""
    return find_duplicate_groups_optimized(entities, threshold)


def rank_group(group, appt_map, ins_map, branch_map):
    """Returns group members sorted best→worst (KEEP first)."""
    ranked = []
    for e in group:
        a = appt_map.get(e["id"], {})
        ranked.append({
            "entity": e,
            "appointments": a.get("total", 0),
            "insurance_mappings": len(ins_map.get(e["id"], [])),
            "branches": len(branch_map.get(e["id"], [])),
        })
    ranked.sort(key=lambda x: (
        -x["appointments"],
        -x["insurance_mappings"],
        -x["branches"],
    ))
    return ranked


# ---------------------------------------------------------------------------
# Terminal output
# ---------------------------------------------------------------------------

DIVIDER = "=" * 78


def section(title: str):
    print(f"\n{DIVIDER}\n  {title}\n{DIVIDER}")


def print_group(group, appt_map, ins_map, branch_map, threshold):
    """Print detailed analysis for a duplicate group."""
    # Similarity scores
    print("\nSimilarity scores (pairs that meet threshold):")
    pairs = []
    for a, b in combinations(group, 2):
        score = similarity(a["name"], b["name"])
        if score >= threshold:
            pairs.append([a["name"], b["name"], f"{round(score)}%"])
    print(tabulate(pairs, headers=["Entity A", "Entity B", "Score"], tablefmt="rounded_outline"))

    # 1. Entity details
    section("1. ENTITY DETAILS")
    rows = [[
        e["id"], e["name"], e["entity_type"] or "—", e["status"],
        e["client_key"] or "—", "Yes" if e["is_onboarded"] else "No",
        e["email"] or "—", e["phone_number"] or "—",
    ] for e in group]
    print(tabulate(rows, tablefmt="rounded_outline",
                   headers=["ID", "Name", "Type", "Status", "Client Key", "Onboarded", "Email", "Phone"]))

    # 2. Appointments
    section("2. APPOINTMENTS")
    rows = []
    for e in group:
        a = appt_map.get(e["id"], {})
        rows.append([e["name"], e["client_key"] or "—",
                     a.get("total", 0), a.get("completed", 0),
                     a.get("pending", 0), a.get("other", 0)])
    print(tabulate(rows, tablefmt="rounded_outline",
                   headers=["Entity", "Client Key", "Total", "Completed", "Pending", "Other"]))

    # 3. Insurance mappings
    section("3. INSURANCE MAPPINGS (entity_to_entity)")
    rows = []
    for e in group:
        mappings = ins_map.get(e["id"], [])
        if not mappings:
            rows.append([e["name"], e["client_key"] or "—", 0, "—", "—", "—"])
        for m in mappings:
            rows.append([e["name"], e["client_key"] or "—", len(mappings),
                         m["insurance_name"], m["mapping_status"],
                         "Yes" if m["can_book_valuation"] else "No"])
    print(tabulate(rows, tablefmt="rounded_outline",
                   headers=["Entity", "Client Key", "Total Mappings", "Insurance",
                             "Mapping Status", "Can Book Val."]))

    # 4. Branches
    section("4. BRANCHES")
    rows = []
    for e in group:
        branches = branch_map.get(e["id"], [])
        if not branches:
            rows.append([e["name"], e["client_key"] or "—", 0, "—", "—", "—", "—"])
        for b in branches:
            rows.append([e["name"], e["client_key"] or "—", len(branches),
                         b["admin_area"] or "—", b["status"] or "—",
                         b["email"] or "—", b["phone_number"] or "—"])
    print(tabulate(rows, tablefmt="rounded_outline",
                   headers=["Entity", "Client Key", "Total Branches",
                             "Admin Area", "Status", "Email", "Phone"]))

    # 5. Summary
    section("5. CONSOLIDATED SUMMARY & REMOVAL SUGGESTION")
    ranked = rank_group(group, appt_map, ins_map, branch_map)
    rows = []
    for i, r in enumerate(ranked):
        e = r["entity"]
        rows.append([e["name"], e["id"], e["client_key"] or "—", e["status"],
                     r["appointments"], r["insurance_mappings"], r["branches"],
                     "KEEP" if i == 0 else "REMOVE"])
    print(tabulate(rows, tablefmt="rounded_outline",
                   headers=["Name", "ID", "Client Key", "Status",
                             "Appointments", "Insurance Mappings", "Branches", "Action"]))

    keep = ranked[0]["entity"]
    print(f"\n  KEEP   → {keep['name']} (client_key={keep['client_key']}, id={keep['id']})")
    for r in ranked[1:]:
        e = r["entity"]
        print(f"  REMOVE → {e['name']} (client_key={e['client_key']}, id={e['id']})")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

# Palette
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
_KEEP_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green
_REM_FILL  = PatternFill("solid", fgColor="FFDDC1")   # light orange
_HDR_FONT  = Font(bold=True, color="FFFFFF")
_BOLD       = Font(bold=True)
_THIN       = Side(style="thin", color="BFBFBF")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hdr_row(ws, row_num: int, values: list):
    """Create a styled header row."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill   = _HDR_FILL
        c.font   = _HDR_FONT
        c.border = _BORDER
        c.alignment = _CENTER


def _data_cell(ws, row_num: int, col: int, value, fill=None, bold=False, center=False):
    """Create a styled data cell."""
    c = ws.cell(row=row_num, column=col, value=value)
    if fill:
        c.fill = fill
    if bold:
        c.font = _BOLD
    c.border    = _BORDER
    c.alignment = _CENTER if center else _LEFT
    return c


def _auto_width(ws, min_w=10, max_w=50):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            max(min_w, min(length + 3, max_w))


def build_workbook(groups, appt_map, ins_map, branch_map, threshold) -> Workbook:
    """Build complete Excel workbook with all analysis sheets."""
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    _sheet_workplan(wb, groups, appt_map, ins_map, branch_map)
    _sheet_appointments(wb, groups, appt_map)
    _sheet_mappings(wb, groups, ins_map)
    _sheet_branches(wb, groups, branch_map)
    _sheet_similarity(wb, groups, threshold)

    return wb


def _sheet_workplan(wb, groups, appt_map, ins_map, branch_map):  # noqa: PLR0913
    """Create main workplan/summary sheet."""
    ws = wb.create_sheet("Workplan (Summary)")
    ws.freeze_panes = "A2"

    headers = [
        "Group #", "Group Name", "Entity ID", "Entity Name", "Type",
        "Status", "Client Key", "Onboarded", "Email", "Phone",
        "Appointments", "Insurance Mappings", "Branches",
        "Name Similarity %", "Email Match", "Recommended Action",
        # Tracker columns for the team
        "Decision", "Assignee", "Resolved", "Notes",
    ]
    _hdr_row(ws, 1, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for g_idx, group in enumerate(groups, 1):
        ranked = rank_group(group, appt_map, ins_map, branch_map)

        # Best name similarity and whether any pair shares a real email.
        pairs = list(combinations(group, 2))
        best_name_score = max(
            round(similarity(a["name"], b["name"])) for a, b in pairs
        ) if len(group) > 1 else 100
        group_email_match = any(email_match(a, b) for a, b in pairs)

        for i, r in enumerate(ranked):
            e      = r["entity"]
            action = "KEEP" if i == 0 else "REMOVE"
            fill   = _KEEP_FILL if action == "KEEP" else _REM_FILL

            vals = [
                g_idx, group[0]["name"], e["id"], e["name"],
                e["entity_type"] or "", e["status"], e["client_key"] or "",
                "Yes" if e["is_onboarded"] else "No",
                e["email"] or "", e["phone_number"] or "",
                r["appointments"], r["insurance_mappings"], r["branches"],
                f"{best_name_score}%", "Yes" if group_email_match else "", action,
                "", "", "", "",   # tracker blanks
            ]
            for col, val in enumerate(vals, 1):
                _data_cell(ws, row, col, val, fill=fill,
                           bold=(col == 16), center=(col in {1, 11, 12, 13, 14, 15, 16}))
            row += 1

    _auto_width(ws)
    # Lock tracker columns width
    for col_letter in ["P", "Q", "R", "S"]:
        ws.column_dimensions[col_letter].width = 18


def _sheet_appointments(wb, groups, appt_map):
    """Create appointments detail sheet."""
    ws = wb.create_sheet("Appointments")
    ws.freeze_panes = "A2"

    headers = ["Group #", "Group Name", "Entity ID", "Entity Name",
               "Client Key", "Total", "Completed", "Pending", "Other"]
    _hdr_row(ws, 1, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for g_idx, group in enumerate(groups, 1):
        ranked = rank_group(group, appt_map, {}, {})
        for i, r in enumerate(ranked):
            e    = r["entity"]
            a    = appt_map.get(e["id"], {})
            fill = _KEEP_FILL if i == 0 else _REM_FILL
            for col, val in enumerate([
                g_idx, group[0]["name"], e["id"], e["name"], e["client_key"] or "",
                a.get("total", 0), a.get("completed", 0),
                a.get("pending", 0), a.get("other", 0),
            ], 1):
                _data_cell(ws, row, col, val, fill=fill,
                           center=(col in {1, 6, 7, 8, 9}))
            row += 1

    _auto_width(ws)


def _sheet_mappings(wb, groups, ins_map):
    """Create insurance mappings detail sheet."""
    ws = wb.create_sheet("Insurance Mappings")
    ws.freeze_panes = "A2"

    headers = ["Group #", "Group Name", "Entity ID", "Entity Name",
               "Client Key", "Total Mappings", "Insurance Company",
               "Mapping ID", "Mapping Status", "Can Book Valuation",
               "Markup", "Vatable"]
    _hdr_row(ws, 1, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for g_idx, group in enumerate(groups, 1):
        ranked = rank_group(group, {}, ins_map, {})
        for i, r in enumerate(ranked):
            e        = r["entity"]
            mappings = ins_map.get(e["id"], [])
            fill     = _KEEP_FILL if i == 0 else _REM_FILL

            if not mappings:
                for col, val in enumerate([
                    g_idx, group[0]["name"], e["id"], e["name"],
                    e["client_key"] or "", 0, "—", "—", "—", "—", "—", "—",
                ], 1):
                    _data_cell(ws, row, col, val, fill=fill, center=(col in {1, 6}))
                row += 1
            else:
                for m in mappings:
                    for col, val in enumerate([
                        g_idx, group[0]["name"], e["id"], e["name"],
                        e["client_key"] or "", len(mappings),
                        m["insurance_name"], m["mapping_id"], m["mapping_status"],
                        "Yes" if m["can_book_valuation"] else "No",
                        m["markup"], "Yes" if m["vatable"] else "No",
                    ], 1):
                        _data_cell(ws, row, col, val, fill=fill,
                                   center=(col in {1, 6, 9, 10, 11, 12}))
                    row += 1

    _auto_width(ws)


def _sheet_branches(wb, groups, branch_map):
    """Create branches detail sheet."""
    ws = wb.create_sheet("Branches")
    ws.freeze_panes = "A2"

    headers = ["Group #", "Group Name", "Entity ID", "Entity Name",
               "Client Key", "Total Branches", "Branch ID",
               "Admin Area", "Status", "Email", "Phone", "Address"]
    _hdr_row(ws, 1, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for g_idx, group in enumerate(groups, 1):
        ranked = rank_group(group, {}, {}, branch_map)
        for i, r in enumerate(ranked):
            e        = r["entity"]
            branches = branch_map.get(e["id"], [])
            fill     = _KEEP_FILL if i == 0 else _REM_FILL

            if not branches:
                for col, val in enumerate([
                    g_idx, group[0]["name"], e["id"], e["name"],
                    e["client_key"] or "", 0, "—", "—", "—", "—", "—", "—",
                ], 1):
                    _data_cell(ws, row, col, val, fill=fill, center=(col in {1, 6}))
                row += 1
            else:
                for b in branches:
                    for col, val in enumerate([
                        g_idx, group[0]["name"], e["id"], e["name"],
                        e["client_key"] or "", len(branches),
                        b["branch_id"], b["admin_area"] or "—",
                        b["status"] or "—", b["email"] or "—",
                        b["phone_number"] or "—", b["physical_address"] or "—",
                    ], 1):
                        _data_cell(ws, row, col, val, fill=fill, center=(col in {1, 6, 9}))
                    row += 1

    _auto_width(ws)


def _sheet_similarity(wb, groups, threshold):
    """Create similarity scores detail sheet."""
    ws = wb.create_sheet("Similarity Scores")
    ws.freeze_panes = "A2"

    headers = ["Group #", "Group Name", "Entity A", "Entity B",
               "Name Score %", "Email Match", "Meets Threshold"]
    _hdr_row(ws, 1, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    row = 2
    for g_idx, group in enumerate(groups, 1):
        for a, b in combinations(group, 2):
            name_score = round(similarity(a["name"], b["name"]))
            em         = email_match(a, b)
            meets      = entity_similarity(a, b) >= threshold
            fill       = _KEEP_FILL if meets else None
            for col, val in enumerate([
                g_idx, group[0]["name"], a["name"], b["name"],
                f"{name_score}%", "Yes" if em else "", "Yes" if meets else "No",
            ], 1):
                _data_cell(ws, row, col, val, fill=fill, center=(col in {1, 5, 6, 7}))
            row += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--threshold", type=int, default=None,
                        help="Fuzzy similarity threshold %% (default: 85)")
    parser.add_argument("--dsn", default=None,
                        help="PostgreSQL DSN (or set DSN in .env)")
    parser.add_argument("--xlsx", default=None,
                        help="Excel output path (default: duplicate_report_<timestamp>.xlsx)")
    parser.add_argument("--no-terminal", action="store_true",
                        help="Suppress terminal table output (Excel only)")
    parser.add_argument("--config", default=None,
                        help="YAML configuration file path")
    parser.add_argument("--dry-run", action="store_true",
                        help="Analyze without generating Excel")
    parser.add_argument("--debug", action="store_true",
                        help="Enable debug logging")
    args = parser.parse_args()

    # Configure logging level
    if args.debug:
        logger.setLevel(logging.DEBUG)

    # Load configuration
    config = load_config(args.config)
    
    # Use command line args or config values
    threshold = args.threshold or config.get('threshold', 85)
    dsn = args.dsn or config.get('dsn') or os.getenv("DSN")
    if not dsn:
        parser.error("No DSN provided. Set DSN in .env or pass --dsn.")
    
    # Generate output path
    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = config.get('output_dir', '.')
        prefix = config.get('excel_prefix', 'duplicate_report')
        xlsx_path = str(Path(output_dir) / f"{prefix}_{timestamp}.xlsx")

    logger.info(f"Starting duplicate entity analysis")
    logger.info(f"Connected  : {_safe_dsn(dsn)}")
    logger.info(f"Threshold  : {threshold}%")
    if not args.dry_run:
        logger.info(f"Excel out  : {xlsx_path}")

    conn = connect(dsn)
    cur  = conn.cursor()

    try:
        entities   = load_entities(cur)
        appt_map   = load_appointments(cur)
        ins_map    = load_insurance_mappings(cur)
        branch_map = load_branches(cur)

        if not entities:
            logger.warning("No entities loaded from database")
            return

        logger.info(f"Entities   : {len(entities)} loaded — scanning for duplicates...")

        # Find duplicate groups
        groups = find_duplicate_groups_optimized(entities, threshold)

        if not groups:
            logger.info(f"No duplicate groups found at {threshold}% threshold.")
            if not args.dry_run:
                # Still create empty report for documentation
                wb = Workbook()
                ws = wb.active
                ws.title = "No Duplicates Found"
                ws.cell(1, 1, f"No duplicate entities found at {threshold}% threshold")
                ws.cell(2, 1, f"Total entities analyzed: {len(entities)}")
                ws.cell(3, 1, f"Analysis timestamp: {datetime.now().isoformat()}")
                wb.save(xlsx_path)
                logger.info(f"Saved empty report → {xlsx_path}")
            return

        logger.info(f"Found {len(groups)} duplicate group(s).")

        # Terminal output
        if not args.no_terminal:
            for idx, group in enumerate(groups, 1):
                print(f"\n{'#' * 78}")
                print(f"  DUPLICATE GROUP {idx} of {len(groups)}:  {group[0]['name']}")
                print(f"{'#' * 78}")
                print_group(group, appt_map, ins_map, branch_map, threshold)

        # Excel output
        if not args.dry_run:
            logger.info("Building Excel workbook...")
            wb = build_workbook(groups, appt_map, ins_map, branch_map, threshold)
            wb.save(xlsx_path)
            logger.info(f"Saved → {xlsx_path}")
        else:
            logger.info("Dry run completed — no Excel file generated")

    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        sys.exit(1)
    finally:
        cur.close()
        conn.close()
        logger.info(f"\n{DIVIDER}\n  Analysis complete.\n{DIVIDER}")


if __name__ == "__main__":
    main()